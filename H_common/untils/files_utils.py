import csv
import json
import os
import random

from dotenv import load_dotenv
from openpyxl import load_workbook


class JsonUtils:
    def __init__(self):
        self.data = []

    def merge_files(self, file_paths, output_file):
        """
        合并多个 JSON 或 JSONL 文件中的数据，并保存到新文件中。

        Args:
        - file_paths (list): 包含输入文件路径的列表。
        - output_file (str): 输出文件的路径。

        Returns:
        - str: 输出文件的路径。
        """
        for file_path in file_paths:
            if file_path.endswith('.json'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.data.extend(data)
            elif file_path.endswith('.jsonl'):
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        data = json.loads(line)
                        self.data.append(data)
            else:
                raise ValueError(f"Unsupported file format for file: {file_path}. Only .json and .jsonl are supported.")
        return self._save_merged_data(output_file)

    def _save_merged_data(self, output_file):
        """
        将合并后的数据保存到输出文件中。

        Args:
        - output_file (str): 输出文件的路径。

        Returns:
        - str: 输出文件的路径。
        """
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=4)
        return output_file

    def shuffle_data(self):
        """
        对合并后的数据进行随机排序。
        """
        random.shuffle(self.data)

    @staticmethod
    def write_list_to_json(data, path, append=False):
        """
        将列表数据写入 JSON 文件。

        :param data: 要写入的数据列表
        :param path: JSON 文件的路径
        :param append: 是否追加写入，如果为 True，则数据会追加到现有文件末尾；如果为 False，则会覆盖文件内容
        """
        if not isinstance(data, list):
            raise ValueError("data 参数必须是一个列表")
        if append and os.path.exists(path):
            with open(path, 'r') as file:
                try:
                    existing_data = json.load(file)
                    if not isinstance(existing_data, list):
                        raise ValueError("文件中现有的数据不是列表格式")
                except json.JSONDecodeError:
                    existing_data = []
        else:
            existing_data = []
        merged_data = existing_data + data
        with open(path, 'w', encoding='utf-8') as file:
            json.dump(merged_data, file, indent=4, ensure_ascii=False)
        print(f"数据已成功写入到 {path}")

    @staticmethod
    def write_dict_to_json(new_data, file_path):
        """
        将给定的字典数据追加到 JSON 文件中的数组中。

        Parameters:
        - new_data (dict): 要追加的新字典数据。
        - file_path (str): 要追加数据的 JSON 文件路径。
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except FileNotFoundError:
            data = []
        data.append(new_data)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print(f"数据已成功追加到 {file_path}")


class CSVUtils:
    @staticmethod
    def write_data(data, file_path):
        try:
            with open(file_path, mode='a', newline='') as file:  # 打开文件，以追加模式写入
                writer = csv.writer(file)
                for row in data:
                    writer.writerow(row)  # 逐行写入数据，每行末尾自动添加换行符
            print(f"Data successfully appended to {file_path}")
        except IOError as e:
            print(f"Error appending to {file_path}")

    @staticmethod
    def read_csv(file_path):
        data = []
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile)
            for row in csvreader:
                data.append(row)
        return data


class XLSXUtils:
    @staticmethod
    def read_xlsx(file_path):
        """读取xlsx文件并返回对应的数据"""
        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active  # 获取活动工作表
        data = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), 0):  # 从0开始索引
            data.append(row)
        wb.close()
        return data

    @staticmethod
    def get_goods_map(file_path):
        """读取商品数据，返回数据格式：{商品id: 商品数据}"""

        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active  # 获取活动工作表

        map_data = {}
        attribute_name = None
        for idx, row in enumerate(ws.iter_rows(values_only=True), 0):  # 从0开始索引
            if idx not in (0, 1):
                attribute_map = {}
                for idx2, item in enumerate(attribute_name):
                    attribute_map[item] = row[idx2]
                map_data[row[0]] = attribute_map
            elif idx == 1:
                attribute_name = row

        wb.close()
        return map_data


class TxtUtils:

    @staticmethod
    def write_list_to_txt(data, file_path, append=False):
        """
        将列表数据写入文本文件。

        :param data: 要写入的数据列表
        :param file_path: 文本文件的路径
        :param append: 是否追加写入，如果为 True，则数据会追加到现有文件末尾；如果为 False，则会覆盖文件内容
        """
        if not isinstance(data, list):
            raise ValueError("data 参数必须是一个列表")
        mode = 'a' if append else 'w'
        try:
            with open(file_path, mode, encoding='utf-8') as file:
                for item in data:
                    file.write(f"{item}\n")
            print(f"数据已成功写入到 {file_path}")
        except Exception as e:
            print(f"写入文件时发生错误: {e}")


# 示例用法
if __name__ == "__main__":
    load_dotenv()

    data_to_write = [
        ['John Doe', 30],
        ['Jane Smith', 28],
        ['Mike Brown', 35]
    ]

    CSVUtils.write_data(data_to_write, os.getenv('TF-IDF-FILE'))



    # merger = JsonUtils()
    # input_files = [os.getenv('TRAD-MULTI-CHOICE-100'), os.getenv('GOODS_SFT_JSON')]  # 替换成实际的文件路径列表
    #
    # merged_file_path = merger.merge_files(input_files, os.getenv('OUTPUT_FILE'))
    # print(f"合并后的文件已保存到：{merged_file_path}")
    #
    # merger.shuffle_data()
    # print("数据已经被随机排序。")
    #
    # shuffled_file_path = merger._save_merged_data(os.getenv('SHUFFLED_OUTPUT_FILE'))
    # print(f"随机排序后的数据已保存到：{shuffled_file_path}")

    # file_path = 'D:\D盘桌面\软通\袋鼠妈妈项目\微调数据\商品基础信息goods_baseinfo_20240626113925_temp.xlsx'  # 替换为你的XLSX文件路径
    # map_data = XLSXUtils.get_goods_map(file_path)
    # for k in map_data:
    #     print(f"Row : {map_data[k]}")

    # file_path = '商品基础信息goods_baseinfo_20240626113925.xlsx'  # 替换为你的CSV文件路径
    # content = CSVUtils.read_csv(file_path)
    # for row in content:
    #     print(row)